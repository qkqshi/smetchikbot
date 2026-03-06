import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def generate_kp_excel(project_name, items_data, furniture_data=None):
    """
    Генерирует Excel файл с расчетами КП
    
    Args:
        project_name: название проекта
        items_data: список словарей с данными об изделиях
        furniture_data: список словарей с данными о мебели
    
    Returns:
        путь к созданному Excel файлу
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "КП"
    
    # Заголовок
    ws['A1'] = project_name
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='left')
    ws.merge_cells('A1:E1')
    
    # Заголовки таблицы (без заголовков - как на скриншоте)
    row = 2
    
    grand_total_col_c = 0
    grand_total_col_d = 0
    
    # МАТЕРИАЛЫ И МЕБЕЛЬ - совмещенная таблица
    all_items = []
    
    # Добавляем материалы
    for item_data in items_data:
        item = item_data.get('item', {})
        result = item_data.get('result', {})
        total = result.get('total', item_data.get('total_cost', 0))
        quantity = item_data.get('quantity', 1)
        
        all_items.append({
            'name': item.get('name', item_data.get('name', '')),
            'quantity': quantity,
            'price_c': round(total, 1),
            'price_d': round(total, 1)
        })
    
    # Добавляем мебель
    if furniture_data:
        for furn in furniture_data:
            quantity = furn['quantity']
            if isinstance(quantity, str):
                quantity = int(quantity)
            
            price_per_unit = furn['price_per_unit']
            if isinstance(price_per_unit, str):
                price_per_unit = float(price_per_unit.replace(' ', '').replace(',', ''))
            
            total_price = furn['total_price']
            if isinstance(total_price, str):
                total_price = float(total_price.replace(' ', '').replace(',', ''))
            
            all_items.append({
                'name': furn['name'],
                'quantity': quantity,
                'price_c': round(price_per_unit, 1),
                'price_d': round(total_price, 1)
            })
    
    # Выводим все элементы
    for item in all_items:
        # Название (колонка A)
        cell_a = ws.cell(row=row, column=1)
        cell_a.value = item['name']
        cell_a.alignment = Alignment(horizontal='left', vertical='center')
        cell_a.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Количество (колонка B)
        cell_b = ws.cell(row=row, column=2)
        cell_b.value = item['quantity']
        cell_b.alignment = Alignment(horizontal='center', vertical='center')
        cell_b.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Цена 1 (колонка C)
        cell_c = ws.cell(row=row, column=3)
        cell_c.value = item['price_c']
        cell_c.number_format = '#,##0.0'
        cell_c.alignment = Alignment(horizontal='right', vertical='center')
        cell_c.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Цена 2 (колонка D)
        cell_d = ws.cell(row=row, column=4)
        cell_d.value = item['price_d']
        cell_d.number_format = '#,##0.0'
        cell_d.alignment = Alignment(horizontal='right', vertical='center')
        cell_d.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        grand_total_col_c += item['price_c']
        grand_total_col_d += item['price_d']
        row += 1
    
    # Пустая строка
    row += 1
    
    # ИТОГО
    cell_a = ws.cell(row=row, column=1)
    cell_a.value = "Итого"
    cell_a.font = Font(bold=True, size=12)
    cell_a.alignment = Alignment(horizontal='left', vertical='center')
    
    cell_c = ws.cell(row=row, column=3)
    cell_c.value = grand_total_col_c
    cell_c.number_format = '#,##0.0'
    cell_c.font = Font(bold=True, size=12)
    cell_c.alignment = Alignment(horizontal='right', vertical='center')
    cell_c.border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    
    cell_d = ws.cell(row=row, column=4)
    cell_d.value = grand_total_col_d
    cell_d.number_format = '#,##0.0'
    cell_d.font = Font(bold=True, size=12)
    cell_d.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    cell_d.alignment = Alignment(horizontal='right', vertical='center')
    cell_d.border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    
    # Настройка ширины столбцов
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    # Сохраняем файл
    output_path = f"outputs/{project_name.replace(' ', '_')}.xlsx"
    os.makedirs('outputs', exist_ok=True)
    wb.save(output_path)
    
    return output_path
